from io import BytesIO

from openpyxl import load_workbook


def test_login_success_failed_and_me(client):
    success = client.post(
        "/auth/login",
        json={"email": "admin@example.com", "password": "StrongPassword123!"},
    )
    assert success.status_code == 200
    body = success.json()
    assert body["token_type"] == "bearer"
    assert "password_hash" not in body["user"]

    me = client.get("/auth/me", headers={"Authorization": f"Bearer {body['access_token']}"})
    assert me.status_code == 200
    assert me.json()["email"] == "admin@example.com"

    failed = client.post(
        "/auth/login",
        json={"email": "admin@example.com", "password": "WrongPassword!"},
    )
    assert failed.status_code == 401
    assert failed.json()["error"]["code"] == "INVALID_CREDENTIALS"


def test_role_guards(client, admin_headers, superadmin_headers):
    assert client.get("/users", headers=admin_headers).status_code == 403
    assert client.get("/users", headers=superadmin_headers).status_code == 200
    assert client.get("/employees").status_code == 401

    created = client.post(
        "/users",
        headers=superadmin_headers,
        json={
            "name": "Admin Baru",
            "email": "baru@example.com",
            "password": "AnotherStrong123!",
            "role": "admin",
            "is_active": True,
        },
    )
    assert created.status_code == 201
    assert "password_hash" not in created.json()
    user_id = created.json()["id"]
    token = client.post(
        "/auth/login",
        json={"email": "baru@example.com", "password": "AnotherStrong123!"},
    ).json()["access_token"]
    assert client.delete(f"/users/{user_id}", headers=superadmin_headers).status_code == 204
    assert client.get("/auth/me", headers={"Authorization": f"Bearer {token}"}).status_code == 401
    inactive_login = client.post(
        "/auth/login",
        json={"email": "baru@example.com", "password": "AnotherStrong123!"},
    )
    assert inactive_login.status_code == 401
    assert inactive_login.json()["error"]["code"] == "USER_INACTIVE"


def test_audit_logs_created_for_crud(client, admin_headers, superadmin_headers):
    employee = client.post(
        "/employees", json={
            "nama": "Audit Employee",
            "jabatan": "Analyst",
            "departemen": "Finance",
            "tanggal_masuk": "2025-01-01",
            "status_aktif": True,
        }, headers=admin_headers,
    ).json()
    client.put(
        f"/employees/{employee['id']}",
        json={**{key: employee[key] for key in ["nama", "jabatan", "departemen", "tanggal_masuk", "status_aktif"]}, "jabatan": "Senior Analyst"},
        headers=admin_headers,
    )
    client.delete(f"/employees/{employee['id']}", headers=admin_headers)

    logs = client.get("/audit-logs?page_size=100", headers=superadmin_headers)
    assert logs.status_code == 200
    actions = {item["action"] for item in logs.json()["items"]}
    assert {"CREATE_EMPLOYEE", "UPDATE_EMPLOYEE", "DELETE_EMPLOYEE"}.issubset(actions)


def test_excel_export_is_valid_and_audited(client, admin_headers, superadmin_headers):
    client.post(
        "/employees", json={
            "nama": "Export Employee",
            "jabatan": "Officer",
            "departemen": "Operations",
            "tanggal_masuk": "2025-01-01",
            "status_aktif": True,
        }, headers=admin_headers,
    )
    response = client.get("/exports/employees.xlsx", headers=admin_headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.active["A1"].value == "ID"
    assert workbook.active.max_row == 2

    logs = client.get("/audit-logs?action=EXPORT_EXCEL", headers=superadmin_headers)
    assert logs.status_code == 200
    assert logs.json()["total"] >= 1
