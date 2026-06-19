from calendar import monthrange
from datetime import date, datetime, time, timezone
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from ..audit import write_audit_log
from ..database import get_db
from ..models import AuditLog, Employee, EmployeeLeave, LeaveType, User
from ..security import require_admin, require_superadmin

router = APIRouter(prefix="/exports", tags=["exports"])


def workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    encoded = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4ED8")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 12), 42)


@router.get("/employees.xlsx")
def export_employees(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    employees = list(db.scalars(select(Employee).order_by(Employee.nama.asc())).all())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Karyawan"
    sheet.append(["ID", "Nama", "Jabatan", "Departemen", "Tanggal Masuk", "Status Aktif", "Dibuat", "Diperbarui"])
    for item in employees:
        sheet.append([
            item.id, item.nama, item.jabatan, item.departemen, item.tanggal_masuk.isoformat(),
            "Aktif" if item.status_aktif else "Tidak Aktif",
            item.created_at.isoformat(), item.updated_at.isoformat(),
        ])
    style_sheet(sheet)
    write_audit_log(
        db, request, "EXPORT_EXCEL", "EMPLOYEE", None, current_user,
        new_values={"row_count": len(employees), "format": "xlsx"},
    )
    db.commit()
    return workbook_response(workbook, f"data-karyawan-{date.today().isoformat()}.xlsx")


@router.get("/employee-leaves.xlsx")
def export_leaves(
    request: Request,
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None, ge=2000, le=2100),
    date_from: date | None = None,
    date_to: date | None = None,
    employee_id: int | None = None,
    leave_type: LeaveType | None = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    statement = select(EmployeeLeave).options(joinedload(EmployeeLeave.employee))
    if month and year:
        period_start = date(year, month, 1)
        period_end = date(year, month, monthrange(year, month)[1])
        statement = statement.where(
            EmployeeLeave.start_date <= period_end,
            EmployeeLeave.end_date >= period_start,
        )
    if date_from:
        statement = statement.where(EmployeeLeave.end_date >= date_from)
    if date_to:
        statement = statement.where(EmployeeLeave.start_date <= date_to)
    if employee_id:
        statement = statement.where(EmployeeLeave.employee_id == employee_id)
    if leave_type:
        statement = statement.where(EmployeeLeave.leave_type == leave_type)
    leaves = list(db.scalars(statement.order_by(EmployeeLeave.start_date.desc())).all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Cuti Karyawan"
    sheet.append(["ID", "Karyawan", "Departemen", "Jenis Cuti", "Tanggal Mulai", "Tanggal Selesai", "Total Hari", "Keterangan"])
    for item in leaves:
        sheet.append([
            item.id, item.employee.nama, item.employee.departemen, item.leave_type.value,
            item.start_date.isoformat(), item.end_date.isoformat(), item.total_days,
            item.description or "",
        ])
    style_sheet(sheet)
    filters = {
        "month": month, "year": year, "date_from": date_from, "date_to": date_to,
        "employee_id": employee_id, "leave_type": leave_type.value if leave_type else None,
    }
    write_audit_log(
        db, request, "EXPORT_EXCEL", "EMPLOYEE_LEAVE", None, current_user,
        new_values={"row_count": len(leaves), "format": "xlsx", "filters": filters},
    )
    db.commit()
    return workbook_response(workbook, f"data-cuti-karyawan-{date.today().isoformat()}.xlsx")


@router.get("/audit-logs.xlsx")
def export_audit_logs(
    request: Request,
    action: str | None = None,
    entity_type: str | None = None,
    user_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    statement = select(AuditLog)
    if action:
        statement = statement.where(AuditLog.action == action)
    if entity_type:
        statement = statement.where(AuditLog.entity_type == entity_type)
    if user_id:
        statement = statement.where(AuditLog.user_id == user_id)
    if date_from:
        statement = statement.where(AuditLog.created_at >= datetime.combine(date_from, time.min, timezone.utc))
    if date_to:
        statement = statement.where(AuditLog.created_at <= datetime.combine(date_to, time.max, timezone.utc))
    logs = list(db.scalars(statement.order_by(AuditLog.created_at.desc())).all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Audit Log"
    sheet.append(["ID", "Waktu", "User", "Email", "Action", "Entity", "Entity ID", "IP Address"])
    for item in logs:
        sheet.append([
            item.id, item.created_at.isoformat(), item.user_name, item.user_email, item.action,
            item.entity_type, item.entity_id or "", item.ip_address or "",
        ])
    style_sheet(sheet)
    write_audit_log(
        db, request, "EXPORT_EXCEL", "AUDIT_LOG", None, current_user,
        new_values={"row_count": len(logs), "format": "xlsx"},
    )
    db.commit()
    return workbook_response(workbook, f"audit-log-{date.today().isoformat()}.xlsx")
